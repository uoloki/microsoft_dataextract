import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from src.common.logger import get_logger

logger = get_logger(__name__)

class ExcelHandler:
    @staticmethod
    def save_to_excel(data_dict, file_path):
        """Save data to an Excel file."""
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, data in data_dict.items():
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ExcelHandler.adjust_column_width(writer.book[sheet_name])
            logger.info(f"Data saved to {file_path}")
        except Exception as e:
            logger.error(f"Error saving data to Excel: {e}")
            raise

    @staticmethod
    def adjust_column_width(sheet):
        """Adjust the column width of the Excel sheet to fit the content."""
        try:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            logger.info(f"Adjusted column widths for sheet: {sheet.title}")
        except Exception as e:
            logger.error(f"Error adjusting column widths: {e}")
            raise

    @staticmethod
    def load_data(file_path, sheet_name):
        """Load data from an Excel sheet."""
        try:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception as e:
            logger.error(f"Error loading data from Excel: {e}")
            raise

    @staticmethod
    def filter_columns_with_Y(df):
        """Filter columns based on 'Y' in corresponding _Y columns."""
        try:
            columns_to_keep = [col[:-2] for col in df.columns if col.endswith('_Y') and 'Y' in df[col].values]
            filtered_df = df[columns_to_keep]
            return filtered_df
        except Exception as e:
            logger.error(f"Error filtering columns: {e}")
            return df

    @staticmethod
    def process_excel_file(input_file, output_file, processor):
        """Process an Excel file, filtering data and adjusting column widths."""
        try:
            excel_data = pd.ExcelFile(input_file)

            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for sheet_name in excel_data.sheet_names:
                    df = pd.read_excel(excel_data, sheet_name=sheet_name)
                    filtered_df = ExcelHandler.filter_columns_with_Y(df)
                    filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = load_workbook(output_file)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                ExcelHandler.adjust_column_width(sheet)
            workbook.save(output_file)

            logger.info(f"Filtered data has been written to '{output_file}'")
        except Exception as e:
            logger.error(f"An error occurred while processing the Excel file: {e}")
            raise

    @staticmethod
    def add_y_columns(df):
        """Add a 'Y' column for each existing column."""
        for column in df.columns:
            df[f'{column}_Y'] = 'Y'
        return df

    @staticmethod
    def load_and_filter_excel(input_filename, output_filename):
        """Load data from Excel, filter based on _Y columns, and save to new Excel file."""
        try:
            workbook = load_workbook(input_filename)

            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                for sheetname in workbook.sheetnames:
                    sheet = workbook[sheetname]
                    # Load data into a DataFrame
                    data = pd.DataFrame(sheet.values)
                    data.columns = data.iloc[0]
                    data = data[1:]

                    # Filter the data
                    filtered_data = ExcelHandler.filter_columns_with_Y(data)

                    # Save the filtered data to the new Excel file
                    filtered_data.to_excel(writer, sheet_name=sheetname, index=False)

                    # Adjust column widths
                    ExcelHandler.adjust_column_width(writer.sheets[sheetname])

            logger.info(f"Filtered data saved to {output_filename}")
        except Exception as e:
            logger.error(f"Error processing and saving workbook: {e}")
            raise