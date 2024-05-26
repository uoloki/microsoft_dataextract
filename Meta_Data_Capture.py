import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def load_data(file_path, sheet_name):
    """Load data from an Excel sheet."""
    return pd.read_excel(file_path, sheet_name=sheet_name)

def filter_data(df):
    """Filter data based on 'Y' in corresponding _Y columns."""
    filtered_df = pd.DataFrame()
    for column in df.columns:
        if column.endswith('_Y') and df[column].iloc[0] == 'Y':
            original_column = column[:-2]
            filtered_df[original_column] = df[original_column]
    return filtered_df

def save_filtered_data(data_dict):
    """Save filtered data to a new Excel file."""
    file_path = 'filtered_metadata.xlsx'
    with pd.ExcelWriter(file_path) as writer:
        for sheet_name, data in data_dict.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)
            adjust_column_width(file_path, sheet_name)

    print(f"Filtered data saved to {file_path}")

def adjust_column_width(file_path, sheet_name):
    """Adjust the column width of the Excel sheet to fit the content."""
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width
    workbook.save(file_path)

def main():
    """Main function to capture filtered data from Excel."""
    file_path = 'all_metadata.xlsx'
    sheet_names = ['Users', 'Groups', 'Teams', 'Channels', 'Messages', 'Sites', 'Files']

    data_dict = {}
    for sheet_name in sheet_names:
        # Load data from Excel sheet
        df = load_data(file_path, sheet_name)

        # Filter data based on 'Y' in _Y columns
        filtered_df = filter_data(df)

        # Store filtered data
        data_dict[sheet_name] = filtered_df

    # Save filtered data to a new Excel file
    save_filtered_data(data_dict)

if __name__ == "__main__":
    main()
